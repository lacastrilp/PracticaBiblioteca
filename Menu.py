# Importa la clase AVL y la clase Libro del módulo Arbol
from Arbol import AVL, Libro
# Importa la librería pandas con el alias pd
import pandas as pd
# Importa el módulo openpyxl
import openpyxl
# Importa la función load_workbook del módulo openpyxl
from openpyxl import (load_workbook)



# Define una función para mostrar el menú de opciones
def mostrar_menu():
    # Imprime el menú
    print("\nMenú:")
    print("1. Buscar libro por ID")
    print("2. Buscar libro por título")
    print("3. Buscar libro por autor")
    print("4. Prestar libro")
    print("5. Devolver libro")
    print("6. Mostrar todos los libros")
    print("7. Recorrer libros en pre-orden")
    print("8. Recorrer libros en in-orden")
    print("9. Recorrer libros en post-orden")
    print("10. Verificar si el árbol está vacío")
    print("11. Eliminar libro por título")
    print("12. Insertar libro")
    print("13. Buscar libro por ID con BFS")
    print("14. Buscar libro por ID con DFS")
    print("15. Buscar libros disponibles")
    print("16. Imprimir arbol")
    print("17. Salir")

# Define una función recursiva para buscar libros disponibles en el árbol AVL
def buscar_libros_disponibles_recursivo(nodo, disponibles):
    # Si el nodo actual no es nulo (es decir, si existe)
    if nodo:
        # Busca en el subárbol izquierdo
        buscar_libros_disponibles_recursivo(nodo.izquierda, disponibles)
        # Si el estado del libro en el nodo actual es "disponible"
        if nodo.libro.estadoAct == "disponible":
            # Agrega el libro a la lista de libros disponibles
            disponibles.append(nodo.libro)
        # Busca en el subárbol derecho
        buscar_libros_disponibles_recursivo(nodo.derecha, disponibles)

# Define una función para buscar y mostrar los libros disponibles en el árbol AVL
def buscar_libros_disponibles(arbol):
    disponibles = []  # Lista para almacenar los libros disponibles
    # Llama a la función recursiva para buscar libros disponibles
    buscar_libros_disponibles_recursivo(arbol.raiz, disponibles)
    # Si se encontraron libros disponibles
    if disponibles:
        # Imprime los detalles de los libros disponibles
        print("\nLibros disponibles:")
        for libro in disponibles:
            print(f'ID: {libro.id}, \n Título: {libro.titulo}, \n Autor: {libro.autor}\n')
    else:
        # Si no se encontraron libros disponibles, imprime un mensaje
        print("No hay libros disponibles.")

# Define una función para recorrer los libros en pre-orden en el árbol AVL
def recorrer_libros_preorden(arbol):
    # Imprime un mensaje indicando el tipo de recorrido
    print("\nRecorrido pre-orden:\n")
    # Llama al método recorrer_preorden del árbol AVL para realizar el recorrido
    arbol.recorrer_preorden(arbol.raiz)

# Define una función para recorrer los libros en in-orden en el árbol AVL
def recorrer_libros_inorden(arbol):
    # Imprime un mensaje indicando el tipo de recorrido
    print("\nRecorrido in-orden:\n")
    # Llama al método recorrer_inorden del árbol AVL para realizar el recorrido
    arbol.recorrer_inorden(arbol.raiz)

# Define una función para recorrer los libros en post-orden en el árbol AVL
def recorrer_libros_postorden(arbol):
    # Imprime un mensaje indicando el tipo de recorrido
    print("\nRecorrido post-orden:\n")
    # Llama al método recorrer_postorden del árbol AVL para realizar el recorrido
    arbol.recorrer_postorden(arbol.raiz)

# Define una función para verificar si el árbol está vacío
def verificar_arbol_vacio(arbol):
    # Si el árbol está vacío, imprime un mensaje
    if arbol.esta_vacio():
        print("El árbol está vacío.")
    # Si el árbol no está vacío, imprime otro mensaje
    else:
        print("El árbol no está vacío.")
        print(f"Altura del árbol: {arbol.altura(arbol.raiz)}")
        print(f"Cantidad de nodos: {arbol.contar_nodos(arbol.raiz)}")
        print(f"Factor de balance: {arbol.factor_balance(arbol.raiz)}")
        
# Define una función para buscar un libro por su ID en el árbol AVL
def buscar_libro_por_id(arbol):
    
        while True:
            try:
                # Solicita al usuario que ingrese el ID del libro a buscar
                id = int(input("Ingrese el ID del libro: "))
                # Busca el libro en el árbol AVL
                libro = arbol.buscar(arbol.raiz, id)
                # Si el libro se encuentra, imprime sus detalles
                if libro:
                    print(f"\nLibro encontrado: \n Título: {libro.titulo},\n Autor: {libro.autor},\n Estado: {libro.estadoAct}\n")
                #Si el libro no se encuentra, imprime un mensaje indicando que no se encontró
                else:
                    print("Libro no encontrado.")
            except ValueError:
                print("Por favor, ingrese un ID válido.")

            # Preguntar al usuario si desea buscar otro libro
            respuesta = input("¿Desea buscar otro libro? (s/n): ").strip().lower()
            if respuesta != "s":
                break

# Define una función para buscar libros por su título en el árbol AVL
def buscar_libro_por_titulo(arbol):
        while True:
            # Solicita al usuario que ingrese el título del libro a buscar
            texto = input("Ingrese un título de libro: ").strip()
            # Busca los libros por título en el árbol AVL
            arbol.buscar_por_titulo(texto)
            # Preguntar al usuario si desea realizar otra búsqueda
            respuesta = input("¿Desea buscar otro libro por autor? (s/n): ").strip().lower()
            if respuesta != "s":
                break

# Define una función para buscar libros por su autor en el árbol AVL
def buscar_libro_por_autor(arbol):
     while True:
        # Solicita al usuario que ingrese el autor del libro a buscar
        autor = input("Ingrese el autor del libro: ").strip()
        # Busca los libros por autor en el árbol AVL
        arbol.buscar_por_autor(autor.lower())
        # Preguntar al usuario si desea realizar otra búsqueda
        respuesta = input("¿Desea buscar otro libro por autor? (s/n): ").strip().lower()
        if respuesta != "s":
            break

# Define una función para insertar un nuevo libro en el árbol AVL y en el archivo Excel
def insertar_libro(arbol, archivo):
    #   Solicita al usuario que ingrese los detalles del nuevo libro
    while True:
        # try manejador de errores
        try:
            # Solicita al usuario que ingrese los detalles del nuevo libro
            id = int(input("Ingrese el ID del nuevo libro: "))
            titulo = input("Ingrese el título del nuevo libro: ")
            autor = input("Ingrese el autor del nuevo libro: ")
            
            # Verificar si ya existe un libro con el mismo ID, título y autor
            if arbol.buscar(arbol.raiz, id):
                print("Ya existe un libro con el mismo ID en el árbol.")
            elif arbol.buscar_libro_por_titulo_autor(titulo, autor):
                print("Ya existe un libro con el mismo título y autor en el árbol.")
            else:
                # Crear una instancia de Libro con los detalles proporcionados
                libro = Libro(id, titulo, autor, "disponible")
                
                # Insertar el nuevo libro en el árbol AVL
                arbol.raiz = arbol.insertar(arbol.raiz, libro)

                # Abrir el archivo Excel y obtener la hoja activa
                workbook = load_workbook(filename=archivo)
                sheet = workbook.active

                # Verificar si ya existe un libro con el mismo ID en el archivo Excel
                if id in [row[0].value for row in sheet.iter_rows(min_row=2)]:
                    print("Ya existe un libro con el mismo ID en el archivo Excel.")
                else:
                    # Agregar los detalles del nuevo libro como una nueva fila al final del archivo Excel
                    sheet.append([id, titulo, autor, "disponible"])

                    # Guardar los cambios y cerrar el archivo Excel
                    workbook.save(filename=archivo)

                    # Imprimir un mensaje indicando que el libro fue insertado correctamente
                    print(f"Libro '{titulo}' insertado correctamente.")
            
            # Preguntar al usuario si desea realizar otra inserción
            respuesta = input("¿Desea insertar otro libro? (s/n): ").strip().lower()
            if respuesta != "s":
                break
        except ValueError:
            # Imprimir un mensaje si se produce un error al insertar el libro
            print("Por favor, ingrese un ID válido.")
        except Exception as e:
            print(f"Error: {e}")

# Define una función para buscar un libro por su ID en el árbol AVL utilizando BFS
def searchInBFS(arbol, id):
    # Verifica si la entrada es un número
    if id <0 :
        print("Por favor, ingrese un número válido.")
        return

    try:
        # Convierte la entrada a un entero
        id = int(id)

        # Si el árbol está vacío, imprime un mensaje y termina la función
        if not arbol.raiz:
            print("No hay libros en el árbol")
            return

        # Inicializa una cola con el nodo raíz
        cola = [arbol.raiz]
        # Mientras la cola no esté vacía...
        while cola:
            # ... quita el primer nodo de la cola
            x = cola.pop(0)
            if x:
                # Si el id del libro de este nodo coincide con el id buscado...
                if x.libro.id == id:
                    # ... imprime el título del libro y termina la función
                    print(f"Libro encontrado: {x.libro.titulo}")
                    print('autor: {x.libro.autor}')
                    print('estado: {x.libro.estadoAct}')
                    print('su altura es: {x.altura}')
                    
      
                    return
                # Agrega los hijos de este nodo a la cola, si existen
                cola.append(x.izquierda)
                cola.append(x.derecha)

        # Si se ha recorrido todo el árbol y no se ha encontrado el libro, imprime un mensaje
        print("No se encontró un libro con ese ID.")
    except ValueError:
        # Captura la excepción ValueError cuando la entrada no es un entero válido
        print("Por favor, ingrese un ID válido.")

# Define una función para buscar un libro por su ID en el árbol AVL utilizando DFS
def searchInDFS(self, id):
    try:
        # Convierte la entrada a un entero
        id = int(id)

        # Si el árbol está vacío, imprime un mensaje y termina la función
        if not self.raiz:
            print("No hay libros en el árbol")
            return

        # Inicializa una pila con el nodo raíz
        pila = [self.raiz]
        # Mientras la pila no esté vacía...
        while pila:
            # ... quita el último nodo de la pila
            x = pila.pop()
            if x:
                # Si el id del libro de este nodo coincide con el id buscado...
                if x.libro.id == id:
                    # ... imprime los detalles del libro y termina la función
                    print(f"Libro encontrado: {x.libro.titulo}")
                    print(f'Autor: {x.libro.autor}')
                    print(f'Estado: {x.libro.estadoAct}')
                    print(f'Su altura es: {x.altura}')
                    return
                # Agrega los hijos de este nodo a la pila, si existen
                if x.derecha:
                    pila.append(x.derecha)
                if x.izquierda:
                    pila.append(x.izquierda)

        # Si se ha recorrido todo el árbol y no se ha encontrado el libro, imprime un mensaje
        print("No se encontró un libro con ese ID.")
    except ValueError:
        # Captura la excepción ValueError cuando la entrada no es un entero válido
        print("Por favor, ingrese un ID válido (número entero positivo).")

# Define una función para eliminar un libro por su título del árbol AVL y del archivo Excel
def eliminar_libro_por_titulo(arbol, archivo):
    while True:
        try:
            # Solicita al usuario que ingrese el título del libro que desea eliminar
            titulo = input("Ingrese el título del libro que desea eliminar: ")
            # Si el título existe en el índice de títulos del árbol AVL...
            if titulo in arbol.indice_titulo:
                # Obtiene la primera instancia del libro con el título dado (asumiendo que cada título es único)
                libro = arbol.indice_titulo[titulo][0]
                # Elimina el libro del árbol AVL
                arbol.raiz = arbol.eliminar(arbol.raiz, libro.id)
                # Elimina el título del índice de títulos del árbol
                del arbol.indice_titulo[titulo]

                # Lee el archivo Excel en un DataFrame
                df = pd.read_excel(archivo)

                # Elimina la fila correspondiente al libro del DataFrame
                df = df[df['titulo'] != titulo]

                # Escribe el DataFrame resultante de nuevo al archivo Excel sin el libro eliminado
                df.to_excel(archivo, index=False)

                # Imprime un mensaje indicando que el libro fue eliminado correctamente
                print(f"Libro '{titulo}' eliminado correctamente.")
            # Si el título no existe en el índice de títulos del árbol AVL, imprime un mensaje
            else:
                print("Libro no encontrado.")
            # Preguntar al usuario si desea realizar otra búsqueda    
            respuesta = input("¿Desea buscar otro libro por autor? (s/n): ").strip().lower()
            if respuesta != "s":
                break
        except ValueError:
            print("Por favor, ingrese un título válido.")

# Define una función para prestar un libro y actualizar su estado en el archivo Excel
def prestar_libro(arbol, archivo):
    while True:
        try:    
            # Solicita al usuario que ingrese el ID del libro que desea prestar
            id = int(input("Ingrese el ID del libro que desea prestar: "))
            # Presta el libro utilizando la función prestar_libro del árbol AVL
            arbol.prestar_libro(arbol.raiz, id)

            # Abre el archivo Excel y obtiene la hoja activa
            workbook = load_workbook(filename=archivo)
            sheet = workbook.active

            # Encuentra la fila correspondiente al ID del libro en el archivo Excel
            for row in sheet.iter_rows(min_row=2):  # Suponiendo que la primera fila contiene encabezados
                if row[0].value == id:
                    # Modifica la cuarta columna de la fila correspondiente para actualizar el estado del libro a "reservado"
                    row[3].value = "reservado"  # Reemplaza "Nuevo valor" con el valor que quieras poner

            # Guarda los cambios y cierra el archivo Excel
            workbook.save(filename=archivo)
            # Preguntar al usuario si desea realizar otra búsqueda
            respuesta = input("¿Desea buscar otro libro por autor? (s/n): ").strip().lower()
            if respuesta != "s":
                break
        except ValueError:
            print("Por favor, ingrese un ID válido.")

# Define una función para devolver un libro y actualizar su estado en el archivo Excel
def devolver_libro(arbol, archivo):
    while True:
        try:
            # Solicita al usuario que ingrese el ID del libro que desea devolver
            id = int(input("Ingrese el ID del libro que desea devolver: "))
            # Devuelve el libro utilizando la función devolver_libro del árbol AVL
            arbol.devolver_libro(arbol.raiz, id)

            # Abre el archivo Excel y obtiene la hoja activa
            workbook = load_workbook(filename=archivo)
            sheet = workbook.active
            
            # Encuentra la fila correspondiente al ID del libro en el archivo Excel
            for row in sheet.iter_rows(min_row=2):  # Suponiendo que la primera fila contiene encabezados
                if row[0].value == id:
                    # Modifica la cuarta columna de la fila correspondiente para actualizar el estado del libro a "disponible"
                    row[3].value = "disponible"  # Reemplaza "Nuevo valor" con el valor que quieras poner

            # Guarda los cambios y cierra el archivo Excel
            workbook.save(filename=archivo)
            # Preguntar al usuario si desea realizar otra búsqueda
            respuesta = input("¿Desea buscar otro libro por autor? (s/n): ").strip().lower()
            if respuesta != "s":
                break
        except ValueError:
            print("Por favor, ingrese un ID válido.")

# Define una función para mostrar todos los libros en orden
def mostrar_todos_los_libros(arbol):
    print("\nTodos los libros:")
    arbol.recorrer_inorden(arbol.raiz)

# Crear árbol y cargar datos desde un archivo Excel
arbol = AVL()
arbol.cargar_libros_desde_excel('datapijsonn.xlsx')

# Menú principal
while True:
    mostrar_menu()  # Muestra el menú de opciones
    opcion = input("Seleccione una opción: ")  # Solicita al usuario que seleccione una opción
    if opcion == '1':
        buscar_libro_por_id(arbol)  # Llama a la función para buscar un libro por ID
    elif opcion == '2':
        buscar_libro_por_titulo(arbol)  # Llama a la función para buscar un libro por título
    elif opcion == '3':
        buscar_libro_por_autor(arbol)  # Llama a la función para buscar un libro por autor
    elif opcion == '4':
        prestar_libro(arbol, 'datapijsonn.xlsx')  # Llama a la función para prestar un libro
    elif opcion == '5':
        devolver_libro(arbol, 'datapijsonn.xlsx')  # Llama a la función para devolver un libro
    elif opcion == '6':
        mostrar_todos_los_libros(arbol)  # Llama a la función para mostrar todos los libros
    elif opcion == '7':
        recorrer_libros_preorden(arbol)  # Llama a la función para recorrer los libros en pre-orden
    elif opcion == '8':
        recorrer_libros_inorden(arbol)  # Llama a la función para recorrer los libros en in-orden
    elif opcion == '9':
        recorrer_libros_postorden(arbol)  # Llama a la función para recorrer los libros en post-orden
    elif opcion == '10':
        verificar_arbol_vacio(arbol)  # Llama a la función para verificar si el árbol está vacío
    elif opcion == '11':
        eliminar_libro_por_titulo(arbol, 'datapijsonn.xlsx')  # Llama a la función para eliminar un libro por título
        arbol.plot_tree()  # Muestra el árbol AVL actualizado después de la eliminación
    elif opcion == '12':
        insertar_libro(arbol, 'datapijsonn.xlsx')  # Llama a la función para insertar un nuevo libro
        arbol.plot_tree()  # Muestra el árbol AVL actualizado después de la inserción
    elif opcion == '13':
        try:
            id = int(input("Ingrese el ID del libro que desea buscar: "))
            searchInBFS(arbol, id)  # Llama a la función para buscar un libro por ID utilizando BFS
        except ValueError:
            print("Por favor, ingrese un ID válido.")
    elif opcion == '14':
        try:
            id = int(input("Ingrese el ID del libro que desea buscar: "))
            searchInDFS(arbol, id)  # Llama a la función para buscar un libro por ID utilizando DFS
        except ValueError:
            print("Por favor, ingrese un ID válido.")
    elif opcion == '15':
        buscar_libros_disponibles(arbol)  # Llama a la función para buscar libros disponibles
    elif opcion == '16':
        arbol.plot_tree()  # Muestra el árbol AVL
    elif opcion == '17':
        print("¡Hasta luego!")  # Imprime un mensaje de despedida
        break  # Sale del bucle while y termina el programa
    else:
        print("Opción no válida. Por favor, seleccione una opción válida.")  # Imprime un mensaje si la opción no es válida
